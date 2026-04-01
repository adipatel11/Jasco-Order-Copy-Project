import openpyxl
from openpyxl.styles import NamedStyle, Font

order_file_path = "orders.xlsx"
wb = openpyxl.load_workbook(order_file_path)
ws1 = wb["Bulk"]


def writeToFile(row1):

    # row = [item #, name, size, reserved quantity, order #,  date]

    ws1.append(row1)

def deleteDuplicates():

    rows = [list(row) for row in ws1.iter_rows()]

    for row in range(len(rows)):
        newRow = [x.value for x in rows[row]]
        rows[row] = newRow

    rows = list(rows)

    keepGoing = 1
    while(keepGoing):

        keepGoing = 0
        for i in range(len(rows)):
            if rows.count(rows[i]) > 1:
                del rows[i]
                keepGoing = 1
                break
    
            

    ws1.delete_rows(1, ws1.max_row)

    for row in rows:
        ws1.append(row)

    wb.save(order_file_path)
    
    
def cleanFile():

    rows = [list(row) for row in ws1.iter_rows()]

    for row in range(len(rows)):
        newRow = [x.value for x in rows[row]]
        rows[row] = newRow

    rows = list(rows)

    ws1.delete_rows(1, ws1.max_row)

    for row in rows:
        ws1.append(row)
    
    cols = list(ws1.iter_cols())
    dateStyle = NamedStyle(font=Font(name="Calibri", size=12), name="dateStyle", number_format="d-mmm")
    try:
        wb.add_named_style(dateStyle)
    except:
        pass

    for cell in cols[5]:
        cell.style = "dateStyle"

    # setting up font for all columns

    "VLOOKUP(A{row number},SizeData!A$2:B$3839,2,FALSE)"
    
    for num in [0,1,3,4]:
        for cell in cols[num]:
            cell.font = Font(name="Arial", size=12)
    
    cellNum = 0
    for cell in cols[2]:
        cellNum += 1
        if cellNum >= 4:
            cell.font = Font(name="Calibri", size=12)
            cell.value = f"=VLOOKUP(A{cellNum},SizeData!A$2:B$3839,2,FALSE)"

    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    ws1['D1'].value = f'=SUM(D3:D{ws1.max_row+5})'
    ws1['D1'].font = Font(name = "Calibri", size = 16, bold=True)
    for letter in letters:
        ws1[letter + "2"].font = Font(name="Calibri", size = 11)

    wb.save(order_file_path)

"VLOOKUP(A4,SizeData!A$2:B$3839,2,FALSE)"
"=VLOOKUP(A4,SizeData!A$2:B$3839,2,FALSE)"